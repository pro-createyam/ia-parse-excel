"""
mapper.py
----------
Gestion des entêtes Excel, vocabulaire de synonymes,
et détection fuzzy des colonnes.
"""

import re
from typing import Dict, List, Tuple
from rapidfuzz import fuzz

# --- Helpers entêtes/feuilles -----------------------------------------------
def _row_values(ws, row_index: int):
    """
    Retourne les valeurs de la ligne (sans planter si la ligne dépasse les bornes).
    """
    try:
        # openpyxl renvoie une séquence de cellules; si l'index est hors bornes, on retourne vide.
        if row_index < 1 or row_index > getattr(ws, "max_row", 0):
            return []
        return [cell.value for cell in ws[row_index]]
    except Exception:
        return []


def _count_nonempty(vals) -> int:
    """Compte le nombre de cellules non vides (après trim)."""
    n = 0
    for v in vals:
        if v is None:
            continue
        if isinstance(v, str):
            if v.strip() == "":
                continue
        n += 1
    return n


def _is_numeric_like(v) -> bool:
    """
    Heuristique: True si la cellule ressemble à un nombre.
    Accepte virgule décimale, espaces fines, séparateurs de milliers, signes.
    """
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip()
    if s == "":
        return False
    # normalisations soft
    s = s.replace("\u202f", "").replace("\xa0", "").replace(" ", "")
    s = s.replace(",", ".")
    # garde seulement chiffres, '.', '+', '-', éventuellement un seul séparateur décimal
    try:
        float(s)
        return True
    except Exception:
        return False


# Regex date pré-compilées (perf)
_DATE_PATTERNS = (
    re.compile(r"^\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4}$"),  # 31/12/2025, 31-12-25, 31.12.2025
    re.compile(r"^\d{4}[\/\-.]\d{1,2}[\/\-.]\d{1,2}$"),    # 2025/12/31, 2025-12-31, 2025.12.31
)

def _is_date_like_text(s: str) -> bool:
    if not isinstance(s, str):
        return False
    s = s.strip().lower()
    return any(p.match(s) for p in _DATE_PATTERNS)


def _header_keywords_vocab() -> set:
    """
    Construit un vocabulaire de mots d'entêtes (normalisés) à partir de TARGET_SYNONYMS si disponible.
    Ne casse pas si TARGET_SYNONYMS n'est pas encore défini au moment de l'import.
    """
    vocab = set()
    try:
        mapping = TARGET_SYNONYMS  # noqa: F821
        for k, arr in mapping.items():
            vocab.add(_normalize(k))
            for s in arr:
                vocab.add(_normalize(s))
    except Exception:
        # Pas grave: vocab vide => score basé uniquement sur non-vides/textes
        pass
    # Mots génériques utiles même sans TARGET_SYNONYMS
    for kw in (
        "date", "jour", "nom", "prénom", "prenom", "matricule", "service",
        "heures", "hs", "absence", "observations", "contrat", "avenant"
    ):
        vocab.add(_normalize(kw))
    return vocab


def _header_score(vals) -> float:
    """
    Score une ligne pour estimer si c'est une ligne d'entêtes.
    +2 pour chaque cellule texte non vide
    +3 si le texte matche un mot du vocabulaire d'entêtes
    -1 pour chaque cellule numérique
    -1 pour chaque cellule qui ressemble à une date (texte)
    Bonus léger selon # de non-vides.
    """
    vocab = _header_keywords_vocab()
    score = 0.0
    for v in vals:
        if v is None:
            continue
        if isinstance(v, str):
            txt = v.strip()
            if not txt:
                continue
            score += 2.0
            norm = _normalize(txt)
            # match exact ou inclusion simple
            if norm in vocab:
                score += 3.0
            else:
                if any(w in norm for w in vocab):
                    score += 1.5
            if _is_date_like_text(txt):
                score -= 1.0
        else:
            if _is_numeric_like(v):
                score -= 1.0
    nonempty = _count_nonempty(vals)
    score += min(nonempty, 10) * 0.2
    return score


def _best_header_row(ws, max_scan: int = 25) -> int:
    """
    Scanne les 'max_scan' premières lignes et choisit celle qui ressemble le plus
    à une ligne d'en-têtes (score heuristique).
    """
    best_row, best_score = 1, float("-inf")
    last_row = min(getattr(ws, "max_row", 1), max_scan)
    for r in range(1, last_row + 1):
        vals = _row_values(ws, r)
        score = _header_score(vals)
        if score > best_score:
            best_row, best_score = r, score
    return best_row


def _headers_at(ws, header_row_index: int) -> Dict[str, str]:
    """
    Retourne un dict { 'A': 'Nom de colonne', 'B': '...' } sur la ligne d’entête.
    Tolère cellules vides; ignore celles qui le sont.
    Nettoie (trim, remplace retours ligne), garde la chaîne brute sinon.
    """
    from openpyxl.utils import get_column_letter as _gcl  # import local safe
    headers: Dict[str, str] = {}
    try:
        if header_row_index < 1 or header_row_index > getattr(ws, "max_row", 0):
            return headers
        for idx, cell in enumerate(ws[header_row_index], start=1):
            val = getattr(cell, "value", None)
            if val is None:
                continue
            txt = str(val).replace("\n", " ").strip()
            if not txt:
                continue
            headers[_gcl(idx)] = txt
    except Exception:
        # on renvoie ce qui a pu être lu
        return headers
    return headers


def _extract_headers(ws, header_row_index: int) -> Dict[str, str]:
    """
    Récupère les entêtes à partir de la ligne candidate.
    Si très peu d’entêtes détectées, essaie la ligne suivante/précédente et garde le meilleur résultat.
    """
    base = _headers_at(ws, header_row_index)
    if len(base) >= 3:
        return base

    # Essaie r+1 et r-1 pour rattraper les lignes fusionnées/vides
    candidates = [(header_row_index, base)]
    max_row = getattr(ws, "max_row", 1)
    if header_row_index + 1 <= max_row:
        candidates.append((header_row_index + 1, _headers_at(ws, header_row_index + 1)))
    if header_row_index - 1 >= 1:
        candidates.append((header_row_index - 1, _headers_at(ws, header_row_index - 1)))

    # Choisit celui avec le plus d’entrées, puis score d’entête
    def _hdr_quality(item):
        _r, mapping = item
        vals = list(mapping.values())
        return (len(mapping), _header_score(vals))

    _r_best, mapping_best = max(candidates, key=_hdr_quality)
    return mapping_best if mapping_best else base


def _pick_best_sheet(wb):
    """
    Parcourt toutes les feuilles visibles et choisit celle dont la 'meilleure ligne d'entêtes'
    a le score heuristique le plus élevé.
    Retourne (worksheet, header_row_index).
    """
    # point de départ: active si visible
    def _visible_sheets(workbook):
        for ws in workbook.worksheets:
            # sheet_state peut être 'visible', 'hidden', 'veryHidden'
            state = getattr(ws, "sheet_state", "visible")
            if state == "visible":
                yield ws

    visible_ws = list(_visible_sheets(wb)) or [wb.active]

    # init
    best_ws = visible_ws[0]
    best_row_idx = _best_header_row(best_ws)
    best_vals = _row_values(best_ws, best_row_idx)
    best_score = _header_score(best_vals) if best_vals else float("-inf")

    for ws in visible_ws:
        r = _best_header_row(ws)
        vals = _row_values(ws, r)
        sc = _header_score(vals) if vals else float("-inf")
        if sc > best_score:
            best_ws, best_row_idx, best_score = ws, r, sc

    # Garde-fou si la feuille candidate est vide ou quasi vide
    try:
        if getattr(best_ws, "max_row", 0) < 1 or getattr(best_ws, "max_column", 0) < 1:
            return best_ws, 1
    except Exception:
        return best_ws, 1

    return best_ws, best_row_idx


def _normalize(s: str) -> str:
    """
    Normalisation robuste:
      - trim + lower
      - remplace '_' et tirets par espace
      - supprime les accents (NFKD)
      - compacte les espaces multiples (inclut espaces insécables)
    """
    import unicodedata as _ud
    import re as _re
    s = "" if s is None else str(s)
    s = _ud.normalize("NFKD", s)
    s = "".join(c for c in s if not _ud.combining(c))
    # normalisation espaces
    s = s.replace("\u202f", " ").replace("\xa0", " ")
    # séparateurs usuels -> espace
    s = s.replace("_", " ").replace("-", " ")
    s = s.lower().strip()
    s = _re.sub(r"\s+", " ", s)
    return s
# --- Synonymes d'en-têtes attendues (FR/EN/ES + variations usuelles) ---------
TARGET_SYNONYMS = {
    # ────────────────── Identités / clés
    "matricule_salarie": [
        "matricule salarie", "matricule salarié", "matricule", "employee id", "emp id",
        "id salarie", "id salarié", "code salarie", "code salarié", "code agent",
        "badge", "pernr", "personnel number", "matricule rh", "worker id", "staff id",
        "employee number", "num employe", "numéro employé", "numero employe",
    ],
    "matricule_client": [
        "matricule client", "client id", "code client", "code chantier", "code site",
        "site id", "affectation", "code affectation", "cost center", "centre de cout",
        "centre de coût", "centro de costo", "center code", "cost centre",
    ],
    "matricule": [
        "matricule", "employee id", "empid", "id", "code", "code salarie",
        "code salarié", "badge", "pernr", "employee number", "num employe",
    ],
    "cin": [
        "cin", "c.i.n", "id card", "identity", "identity card", "numero cin",
        "num cin", "n cin", "id national", "national id", "piece identite",
        "pièce identité", "dni", "carnet", "national identity",
    ],

    # ────────────────── Référentiel template (identité/contrat)
    "nombre": ["nombre", "nb", "qty", "quantité", "quantite", "quantity", "cantidad"],
    "nom": ["nom", "last name", "surname", "family name", "patronyme", "lastname", "apellidos"],
    "prenom": ["prenom", "prénom", "first name", "given name", "firstname", "nombre"],
    "num_contrat": [
        "n° contrat", "num contrat", "numero contrat", "contract no", "contract number",
        "contrat", "contrat no", "contrat n", "id contrat", "n contrat",
    ],
    "num_avenant": [
        "n° avenant", "num avenant", "avenant", "amendment no", "amendement",
        "avenant no", "avenant n", "contract amendment",
    ],
    "date_debut": [
        "date debut", "date début", "start date", "date debut contrat",
        "debut contrat", "date d debut", "fecha inicio",
    ],
    "date_fin": [
        "date fin", "end date", "date fin contrat", "fin contrat",
        "date d fin", "date de fin", "fecha fin",
    ],
    "service": [
        "service", "departement", "département", "department", "site", "unité", "unite",
        "direction", "pole", "secteur", "equipe", "équipe", "area", "unidad", "departamento",
        "atelier", "workshop", "team",
    ],

    # ────────────────── Jours / compteurs
    "nb_jt": [
        "nb jt", "jours travailles", "jours travaillés", "jours", "nb jours",
        "jours presents", "jours présence", "jours presence", "working days",
        "dias trabajados",
    ],
    "nb_ji": [
        "nb ji", "jours injustifies", "jours injustifiés", "ji", "abs injustifiee",
        "absence injustifiee", "absence injustifiée", "abs non justifiee", "non justifie",
        "unjustified leave", "ausencia injustificada",
    ],
    "nb_cp_280": [
        "280 - nb cp", "cp", "conges payes", "congés payés", "paid leave days",
        "conge paye", "jours cp", "nb cp", "paid leave",
    ],
    "nb_sans_solde": [
        "nb sans solde", "nombre sans solde",
        "sans solde", "conge sans solde", "css", "unpaid leave",
        "conge non paye", "conge non payé", "non paye", "non payé", "unpaid",
    ],
    "nb_jf": [
        "nb jf", "nb jours feries", "nb jours fériés",
        "jours feries", "jours fériés", "public holidays", "jf",
        "jours ferie", "jour ferié", "jours férié", "feriados", "public holiday days",
    ],
    "tx_sal": [
        "tx sal", "taux sal", "taux salarié", "salary rate", "taux horaire", "th",
        "tarif horaire", "rate", "hourly rate",
    ],

    # ────────────────── Heures / primes (codes paie)
    "hrs_norm_010": [
        "010 - hrs norm", "heures normales", "hrs normales", "heure normal", "h. normal",
        "nb heures", "heures", "heures de base", "base hours", "normal hours",
        "regular hours", "horas normales",
    ],
    "rappel_hrs_norm_140": [
        "140 - rappel hrs norm", "rappel heures normales", "rappel 140", "rappel salaire",
        "rappel base", "backpay normal", "regularization normal hours", "regularisation heures normales",
    ],
    "hs_25_020": [
        "020 - hs 25%", "heures sup 25", "hs 25", "maj 25", "hs 25 pourcent",
        "overtime 25", "ot 25", "25%", "o/t 25",
    ],
    "hs_50_030": [
        "030 - hs 50%", "heures sup 50", "hs 50", "maj 50", "hs 50 pourcent",
        "overtime 50", "ot 50", "50%", "o/t 50",
    ],
    "hs_100_050": [
        "050 - hs 100%", "heures sup 100", "hs 100", "maj 100", "hs 100 pourcent",
        "overtime 100", "ot 100", "100%", "o/t 100",
    ],
    "hrs_feries_091": [
        "091 - hrs feries", "heures feries", "heures fériées", "ferie", "férié",
        "jour férié", "heures ferie travaillees", "jf travaille", "holiday worked hours",
        "public holiday hours", "horas feriado trabajadas",
    ],

    "prime_astreinte_462": [
        "462 - prime astreinte", "astreinte", "prime astreinte", "on-call", "on call allowance",
        "oncall", "astreinte prime",
    ],
    "ind_panier_771": [
        "771 - indemn. panier/mois", "panier", "indemnite panier", "prime panier",
        "meal allowance", "panier repas",
    ],
    "ind_transport_777": [
        "777 - ind.transport/mois", "transport", "indemnite transport", "prime transport",
        "transport allowance", "allowance transport",
    ],
    "ind_deplacement_780": [
        "780 - indemnité deplacement", "deplacement", "indemnite deplacement",
        "frais deplacement", "travel allowance", "per diem", "perdiem",
    ],
    "heures_jour_ferie_chome_090": [
        "090 - heures jour ferie chome", "jour ferie chome", "ferie chome", "jf chome",
        "holiday idle hours", "public holiday idle",
    ],

    "observations": [
        "observations", "commentaire", "comments", "notes", "remarques", "note",
        "remarks", "observaciones",
    ],
    "fin_mission": [
        "fin mission", "fin de mission", "end of assignment", "fin contrat mission",
        "mission terminee", "mission terminée", "end of contract", "terminacion de mision",
    ],

    # ────────────────── Timesheet (généraux)
    "date": [
        "date", "jour", "day", "date jour", "work date", "date travail",
        "fecha", "fecha trabajo",
    ],
    "absence": [
        "absence", "motif", "type jour", "statut jour", "am/pm", "am pm",
        "demi journee", "half day", "leave type", "reason", "motivo",
    ],
    "heures_norm": [
        "heures", "heures travaillees", "heures travaillées", "nbr heures",
        "hours worked", "h. normal", "heures jour", "total heures",
        "temps de travail", "duree travail", "durée travail", "worked hours",
        # pour les fichiers où la colonne s'appelle exactement comme ça :
        "heures_norm_dec", "heures norm dec",
    ],
    "hs_25": ["hs 25", "heures sup 25", "maj 25", "overtime 25", "ot 25", "25%", "o/t 25"],
    "hs_50": ["hs 50", "heures sup 50", "maj 50", "overtime 50", "ot 50", "50%", "o/t 50"],
    "hs_100": ["hs 100", "heures sup 100", "maj 100", "overtime 100", "ot 100", "100%", "o/t 100"],
    "hs_feries": [
        "heures feries", "férié", "ferie", "public holiday hours", "holiday worked",
        "jf travaille", "horas feriado",
    ],

    # ────────────────── Colonnes « full name » (utile si le fichier n’a qu’une seule colonne pour nom+prenom)
    "full_name": [
        "nom prenom", "nom et prenom", "nom & prenom", "full name", "employee name",
        "salarié", "salarie", "agent", "collaborateur", "nombre completo", "nombre y apellido",
        "name", "employee",
    ],
}
def _detect_columns(headers: Dict[str, str]) -> Dict[str, str]:
    """
    headers: dict { 'A': 'Intitulé', ... }
    Retourne un mapping { target_key -> column_letter } avec :
      - match exact (après _normalize)
      - détection directe par codes paie (010, 020, 030, 050, 090, 091, 140, 462, 771, 777, 780)
      - fuzzy match via rapidfuzz (token_set_ratio / partial_ratio)
      - résolution des collisions par score (méthode > score)
      - règle spéciale: si 'nom' et 'prenom' tombent sur la même colonne → préférer 'full_name'
    """
    if not headers:
        return {}

    # Vue normalisée des entêtes : {col -> texte normalisé}
    header_norm: Dict[str, str] = {col: _normalize(txt or "") for col, txt in headers.items()}
    header_items: List[Tuple[str, str]] = list(header_norm.items())  # [(col, norm_text), ...]

    # Détection par codes paie (regex à bornes numériques)
    code_to_target = {
        "010": "hrs_norm_010",
        "020": "hs_25_020",
        "030": "hs_50_030",
        "050": "hs_100_050",
        "090": "heures_jour_ferie_chome_090",
        "091": "hrs_feries_091",
        "140": "rappel_hrs_norm_140",
        "462": "prime_astreinte_462",
        "771": "ind_panier_771",
        "777": "ind_transport_777",
        "780": "ind_deplacement_780",
    }
    # Pré-compile les regex de codes avec vraies bornes
    code_regex = {code: re.compile(rf"(?<!\d){code}(?!\d)") for code in code_to_target.keys()}

    # Conteneurs de scores: target -> (col, score, method)
    # method ∈ {"exact", "code", "fuzzy"} ; exact > code > fuzzy à score égal
    ranked: Dict[str, Tuple[str, float, str]] = {}
    # Occupation colonnes: col -> (target, score, method)
    col_taken: Dict[str, Tuple[str, float, str]] = {}

    def _try_assign(target: str, col: str, score: float, method: str):
        """
        Assigne (ou réassigne) target->col si meilleur que l'existant.
        Résout les collisions en gardant la meilleure paire selon (méthode, score).
        """
        if not col:
            return
        prev = ranked.get(target)
        if prev is None or _is_better(method, score, prev[2], prev[1]):
            ranked[target] = (col, score, method)

        # Collision de colonne : si une autre target possède déjà cette colonne, on arbitre
        other = col_taken.get(col)
        if other is None:
            col_taken[col] = (target, score, method)
        else:
            other_target, other_score, other_method = other
            # Qui garde la colonne ?
            if _is_better(method, score, other_method, other_score):
                # La nouvelle paire gagne la colonne
                col_taken[col] = (target, score, method)
                # L'ancienne target perd la colonne si elle pointait dessus
                if ranked.get(other_target, (None, 0, ""))[0] == col:
                    ranked.pop(other_target, None)
            else:
                # L'ancienne paire garde la colonne, on annule la tentative pour cette target si elle y pointait
                if ranked.get(target, (None, 0, ""))[0] == col:
                    ranked.pop(target, None)

    def _is_better(method_a: str, score_a: float, method_b: str, score_b: float) -> bool:
        """Compare (méthode, score). Priorité: exact > code > fuzzy ; à méthode égale, le score tranche."""
        rank = {"exact": 3, "code": 2, "fuzzy": 1}
        ra, rb = rank.get(method_a, 0), rank.get(method_b, 0)
        if ra != rb:
            return ra > rb
        return score_a > score_b

    # 1) Passe exact & codes paie
    for col, raw_norm in header_items:
        # Codes paie explicites
        for code, target in code_to_target.items():
            if code_regex[code].search(raw_norm):
                _try_assign(target, col, 95.0, "code")

        # Match exact vs TARGET_SYNONYMS (ou vs nom de clé)
        for target, syns in TARGET_SYNONYMS.items():
            if target in ranked:  # déjà un match exact pour ce target
                continue
            norm_targets = [_normalize(s) for s in syns] + [_normalize(target)]
            if raw_norm in norm_targets:
                _try_assign(target, col, 100.0, "exact")

    # 2) Passe fuzzy pour les cibles restantes
    for target, syns in TARGET_SYNONYMS.items():
        if target in ranked:
            continue

        syns_norm = [_normalize(s) for s in syns] + [_normalize(target)]
        best_col, best_score = None, 0.0

        for col, htxt in header_items:
            # calcule un score fuzzy robuste : max(token_set, partial)
            try:
                score = max(
                    fuzz.token_set_ratio(htxt, s) for s in syns_norm
                )
                # petit bonus si partial_ratio dépasse aussi un seuil
                pr = max(fuzz.partial_ratio(htxt, s) for s in syns_norm)
                score = max(score, pr * 0.98)  # léger lissage
            except Exception:
                continue

            if score > best_score:
                best_score, best_col = score, col

        if best_col is not None and best_score >= 80:
            _try_assign(target, best_col, float(best_score), "fuzzy")

    # 3) Règles simples de collisions métier
    # - Si même colonne pour nom & prénom => préfère mapper 'full_name'
    nom_col = ranked.get("nom", (None, 0.0, ""))[0]
    prenom_col = ranked.get("prenom", (None, 0.0, ""))[0]
    if nom_col and prenom_col and nom_col == prenom_col:
        # libère nom/prenom et place full_name
        ranked.pop("nom", None)
        ranked.pop("prenom", None)
        # N'affecte pas si full_name existe déjà (et diffère)
        fl = ranked.get("full_name", (None, 0.0, ""))[0]
        if not fl:
            # score médian, méthode 'fuzzy' (ne pas prétendre à un exact)
            _try_assign("full_name", nom_col, 85.0, "fuzzy")

    # Sortie finale: target -> column_letter
    detected: Dict[str, str] = {}
    for target, (col, _score, _method) in ranked.items():
        if col:
            detected[target] = col
    return detected
