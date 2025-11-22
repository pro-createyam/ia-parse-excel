from typing import Optional, List, Any, Dict, Tuple
from openpyxl import load_workbook
from io import BytesIO
import re
import math
from merge import merge_rows
from openpyxl import Workbook
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, Alignment
import json
import calendar
from fastapi import UploadFile, File, Form, HTTPException, Body, Request

from utils_app import _configure_logging, create_app
from utils_data import _parse_rules, _parse_holidays, _coerce_date
from converter import _parse_hours_to_decimal, _parse_days, _hours_to_days, _days_to_hours
from mapper import _normalize, _extract_headers, _pick_best_sheet, _detect_columns

logger = _configure_logging()
app = create_app()

# --- Upload guard ------------------------------------------------------------
MAX_UPLOAD_BYTES = 15 * 1024 * 1024  # 15 MB


# ─────────────────────────── Legacy endpoint (utilisé dans Bubble actuellement)
@app.post("/parse-excel-upload")
async def parse_excel_upload(
    file: UploadFile = File(...),
    holiday_dates: Optional[str] = Form(default=None),
    rules: Optional[str] = Form(default=None),
):
    filename = (getattr(file, "filename", "") or "").lower()
    if not filename or not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        logger.info(
            "parse-excel-upload | rejected filename=%s mime=%s",
            getattr(file, "filename", None), getattr(file, "content_type", None)
        )
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm are supported")
    logger.info(
        "parse-excel-upload | filename=%s mime=%s",
        file.filename, getattr(file, "content_type", None)
    )

    # Lecture + garde-fous taille
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large")

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel: {e}")

    # Choix feuille + entêtes robustes
    ws, header_row_index = _pick_best_sheet(wb)
    headers = _extract_headers(ws, header_row_index)
    if not headers:
        raise HTTPException(status_code=400, detail="No headers detected on the selected sheet.")
    detected = _detect_columns(headers)

    logger.info(
        "parse-excel-upload | sheet=%s header_row=%s headers_sample=%s",
        ws.title, header_row_index, list(headers.values())[:10]
    )

    rules_dict = _parse_rules(rules)
    holidays = _parse_holidays(holiday_dates)

    # Helpers locaux
    def _val_at(r: int, key: str):
        """Lit la valeur à la ligne r pour la clé détectée 'key' (via 'detected')."""
        col = detected.get(key)
        if not col:
            return None
        try:
            return ws[f"{col}{r}"].value
        except Exception:
            return None

    def _hours_at(r: int, col_letter: Optional[str]) -> Optional[float]:
        """Lit et convertit les heures à la ligne r pour la colonne 'col_letter'."""
        if not col_letter:
            return None
        try:
            return _parse_hours_to_decimal(ws[f"{col_letter}{r}"].value)
        except Exception:
            return None

    # Colonnes utiles (fallback entre équivalents)
    COL_MATRICULE = detected.get("matricule") or detected.get("matricule_salarie")
    COL_NOM = detected.get("nom")
    COL_PRENOM = detected.get("prenom")
    COL_FULLNAME = detected.get("full_name")
    COL_CIN = detected.get("cin")
    COL_DATE = detected.get("date")

    # Heures normales : accepte 'heures_norm' (timesheet) ou code paie 'hrs_norm_010'
    COL_HN = (
    detected.get("heures_norm")
    or detected.get("hrs_norm_010")
    or detected.get("heures_norm_dec")
    )
    # Heures sup décomposées
    COL_HS25 = detected.get("hs_25") or detected.get("hs_25_020")
    COL_HS50 = detected.get("hs_50") or detected.get("hs_50_030")
    COL_HS100 = detected.get("hs_100") or detected.get("hs_100_050")
    # Heures fériées (travaillées)
    COL_HFER = detected.get("hs_feries") or detected.get("hrs_feries_091")
    # Absence (pour 'demi_journee' heuristique)
    COL_ABS = detected.get("absence")
    # Jours (compteur direct côté client)
    COL_NBJT = detected.get("nb_jt")
    # Autres compteurs de jours / paramètres paie
    COL_NBJI         = detected.get("nb_ji")
    COL_NBCP_280     = detected.get("nb_cp_280")
    COL_NBSANS_SOLDE = detected.get("nb_sans_solde")
    COL_NBJF         = detected.get("nb_jf")
    COL_TX_SAL       = detected.get("tx_sal")
    COL_RAPPEL140    = detected.get("rappel_hrs_norm_140")
    COL_OBSERVATIONS = detected.get("observations")
    COL_FIN_MISSION  = detected.get("fin_mission")



    rows: List[Dict[str, Any]] = []
    start = max(1, header_row_index + 1)
    max_row = getattr(ws, "max_row", start - 1)
    max_col = getattr(ws, "max_column", 0)
    end = max_row

    # Heuristique de split full name
    def split_full_name(fn: Any) -> Tuple[Optional[str], Optional[str]]:
        """Heuristique : fichiers 'NOM PRENOM' ou 'Prenom Nom' → on sépare."""
        if not fn:
            return None, None
        s = str(fn).strip()
        if not s:
            return None, None
        parts = [p for p in re.split(r"\s+", s) if p]
        if len(parts) == 1:
            return parts[0], None
        prenom = parts[-1]
        nom = " ".join(parts[:-1])
        return nom, prenom

    if end < start:
        # Feuille sans données utiles sous l'entête
        return {
            "rules_used": rules_dict,
            "holiday_dates": holidays,
            "rows": rows,
            "rows_count": 0,
        }

    for r in range(start, end + 1):
        # Identité
        v_matricule = _val_at(r, "matricule") or _val_at(r, "matricule_salarie") or (
            ws[f"{COL_MATRICULE}{r}"].value if COL_MATRICULE else None
        )
        v_cin = _val_at(r, "cin")

        v_nom = _val_at(r, "nom")
        v_prenom = _val_at(r, "prenom")

        if (not v_nom or not v_prenom) and COL_FULLNAME:
            n, p = split_full_name(
                _val_at(r, "full_name") or (ws[f"{COL_FULLNAME}{r}"].value if COL_FULLNAME else None)
            )
            v_nom = v_nom or n
            v_prenom = v_prenom or p

        # Date
        v_date = _coerce_date(
            _val_at(r, "date") or (ws[f"{COL_DATE}{r}"].value if COL_DATE else None)
        )

        # Heures
        h_norm = _hours_at(r, COL_HN)
        hs25 = _hours_at(r, COL_HS25)
        hs50 = _hours_at(r, COL_HS50)
        hs100 = _hours_at(r, COL_HS100)
        hfer = _hours_at(r, COL_HFER)

        hs_normales_agg = None
        parts = [x for x in (hs25, hs50, hs100) if isinstance(x, (int, float)) and math.isfinite(x)]
        if parts:
            hs_normales_agg = round(sum(parts), 2)

        # Jours saisis (nb_jt) si dispo
        nb_jt_val_raw = None
        if COL_NBJT:
            try:
                nb_jt_val_raw = ws[f"{COL_NBJT}{r}"].value
            except Exception:
                nb_jt_val_raw = None
        nb_jt_days = _parse_days(nb_jt_val_raw)
        # Autres compteurs (si présents dans le fichier)
        nb_ji_val = None
        if COL_NBJI:
            try:
                nb_ji_val = _parse_days(ws[f"{COL_NBJI}{r}"].value)
            except Exception:
                nb_ji_val = None

        nb_cp_280_val = None
        if COL_NBCP_280:
            try:
                nb_cp_280_val = _parse_days(ws[f"{COL_NBCP_280}{r}"].value)
            except Exception:
                nb_cp_280_val = None

        nb_sans_solde_val = None
        if COL_NBSANS_SOLDE:
            try:
                nb_sans_solde_val = _parse_days(ws[f"{COL_NBSANS_SOLDE}{r}"].value)
            except Exception:
                nb_sans_solde_val = None

        nb_jf_val = None
        if COL_NBJF:
            try:
                nb_jf_val = _parse_days(ws[f"{COL_NBJF}{r}"].value)
            except Exception:
                nb_jf_val = None

        tx_sal_val = None
        if COL_TX_SAL:
            try:
                tx_sal_val = ws[f"{COL_TX_SAL}{r}"].value
            except Exception:
                tx_sal_val = None

        rappel_hrs_norm_140_val = _hours_at(r, COL_RAPPEL140) if COL_RAPPEL140 else None

        # --- NOUVEAU : Observations + Fin de mission ---
        observations_val = None
        if COL_OBSERVATIONS:
            try:
                observations_val = ws[f"{COL_OBSERVATIONS}{r}"].value
            except Exception:
                observations_val = None

        fin_mission_val = None
        if COL_FIN_MISSION:
            try:
                fin_mission_val = ws[f"{COL_FIN_MISSION}{r}"].value
            except Exception:
                fin_mission_val = None

        # Demi-journée (absence)
        abs_raw = None
        if COL_ABS:
            try:
                abs_raw = ws[f"{COL_ABS}{r}"].value
            except Exception:
                abs_raw = None
        abs_txt = (str(abs_raw).lower().strip()) if abs_raw is not None else ""
        demi_j = True if ("demi" in abs_txt or "1/2" in abs_txt or "half" in abs_txt or abs_txt in {"am", "pm"}) else None

        # Conversions heures ↔ jours
        jours_calcules: Optional[float] = None
        heures_calculees: Optional[float] = None

        # Si uniquement heures → calcule jours
        if (h_norm is not None) and (nb_jt_days is None or nb_jt_days == 0):
            jours_calcules = _hours_to_days(h_norm, rules_dict)

        # Si uniquement jours → calcule heures
        if (nb_jt_days is not None) and (h_norm is None or h_norm == 0):
            heures_calculees = _days_to_hours(nb_jt_days, rules_dict)

        # Fallback demi-journée si rien fourni
        if (h_norm is None) and (nb_jt_days is None) and demi_j:
            jours_calcules = 0.5
            heures_calculees = _days_to_hours(0.5, rules_dict)

        # Raw body (aperçu ligne)
        raw_vals = []
        try:
            for row_tuple in ws.iter_rows(min_col=1, max_col=max_col, min_row=r, max_row=r, values_only=True):
                raw_vals = list(row_tuple)
        except Exception:
            pass
        raw_body_text = " | ".join([str(x) for x in raw_vals if x is not None])[:1000]

        rows.append({
            "matricule": v_matricule,
            "nom": v_nom,
            "prenom": v_prenom,
            "cin": v_cin,
            "date": v_date,

            # heures normales
            "heures_travaillees_decimal": h_norm,
            "heures_norm_dec": h_norm,

            # heures sup détaillées
            "hs_25_dec": hs25,
            "hs_50_dec": hs50,
            "hs_100_dec": hs100,
            "hs_feries_dec": hfer,

            "hs_normales": hs_normales_agg,
            "hs_ferie": hfer,

            # jours
            "nb_jt": nb_jt_days,
            "nb_ji": nb_ji_val,
            "nb_cp_280": nb_cp_280_val,
            "nb_sans_solde": nb_sans_solde_val,
            "nb_jf": nb_jf_val,

            # salaire & rappel
            "tx_sal": tx_sal_val,
            "rappel_hrs_norm_140": rappel_hrs_norm_140_val,

            "jours_calcules": jours_calcules,
            "heures_calculees": heures_calculees,
            "demi_journee": demi_j,
            "raw_body_text": raw_body_text,
            "observations": observations_val,
            "fin_mission": fin_mission_val,

        })

    return {
        "rules_used": rules_dict,
        "holiday_dates": holidays,
        "rows": rows,
        "rows_count": len(rows),
    }


# ─────────────────────────── NEW: Template Intake
@app.post("/template-intake")
async def template_intake(
    file_template: UploadFile = File(...),
    client_id: Optional[str] = Form(default=None),
):
    name = (getattr(file_template, "filename", "") or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        logger.info(
            "template-intake | rejected filename=%s mime=%s client_id=%s",
            getattr(file_template, "filename", None),
            getattr(file_template, "content_type", None),
            client_id,
        )
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm are supported")
    logger.info(
        "template-intake | filename=%s mime=%s client_id=%s",
        file_template.filename, getattr(file_template, "content_type", None), client_id
    )

    # Lecture + garde-fous taille
    try:
        content = await file_template.read()
        if not content:
            raise HTTPException(status_code=400, detail="Empty file")
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail="File too large")
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel: {e}")

    # Choisir la meilleure feuille + ligne d'entêtes (heuristique robuste)
    ws, header_row_index = _pick_best_sheet(wb)
    headers_dict = _extract_headers(ws, header_row_index)
    if not headers_dict:
        raise HTTPException(status_code=400, detail="No headers detected on the selected sheet.")
    logger.info(
        "template-intake | sheet=%s header_row=%s headers_sample=%s",
        ws.title, header_row_index, list(headers_dict.values())[:10]
    )

    # Colonnes attendues côté template paie
    expected = [
        "matricule_salarie","matricule_client","nombre","nom","prenom","num_contrat","num_avenant",
        "date_debut","date_fin","service","nb_jt","nb_ji","nb_cp_280","nb_sans_solde","nb_jf","tx_sal",
        "hrs_norm_010","rappel_hrs_norm_140","hs_25_020","hs_50_030","hs_100_050","hrs_feries_091",
        "prime_astreinte_462","ind_panier_771","ind_transport_777","ind_deplacement_780",
        "heures_jour_ferie_chome_090","observations","fin_mission"
    ]

    # 1) Mapping exact (normalisé) + 2) Fuzzy via _detect_columns
    header_norm: Dict[str, str] = {col: _normalize(txt) for col, txt in headers_dict.items()}
    detected_all = _detect_columns(headers_dict)

    column_map: Dict[str, str] = {}
    # a) ce que la détection a trouvé
    for key in expected:
        if key in detected_all:
            column_map[key] = detected_all[key]

    # b) égalité stricte sur intitulé normalisé (fallback exact)
    for key in expected:
        if key in column_map:
            continue
        key_norm = _normalize(key)
        for col, htxt in header_norm.items():
            if htxt == key_norm:
                column_map[key] = col
                break

    # c) alias de pont (fallback supplémentaire)
    alias_map = {
        "matricule_salarie": ["matricule"],
        "hrs_norm_010":      ["heures_norm"],
        "hs_25_020":         ["hs_25"],
        "hs_50_030":         ["hs_50"],
        "hs_100_050":        ["hs_100"],
        "hrs_feries_091":    ["hs_feries"],
    }
    for target, aliases in alias_map.items():
        if target in column_map:
            continue
        for a in aliases:
            col = detected_all.get(a)
            if col:
                column_map[target] = col
                break

    def _read_cell(col_letter: Optional[str], row: int):
        if not col_letter:
            return None
        try:
            return ws[f"{col_letter}{row}"].value
        except Exception:
            return None

    # Roster (aperçu) : on évite les lignes totalement vides
    roster: List[Dict[str, Any]] = []
    col_matsal = column_map.get("matricule_salarie")
    col_matcli = column_map.get("matricule_client")
    col_nom    = column_map.get("nom")
    col_prenom = column_map.get("prenom")
    col_srv    = column_map.get("service")

    start = header_row_index + 1
    PREVIEW_ROSTER_MAX = 100
    end = min(getattr(ws, "max_row", start - 1), start + PREVIEW_ROSTER_MAX)

    for r in range(start, end + 1):
        row_obj = {
            "row_index_excel": r,
            "matricule_salarie": _read_cell(col_matsal, r),
            "matricule_client": _read_cell(col_matcli, r),
            "nom": _read_cell(col_nom, r),
            "prenom": _read_cell(col_prenom, r),
            "service": _read_cell(col_srv, r),
        }
        # skip si toutes les valeurs affichées sont None/vides
        if all((v is None or (isinstance(v, str) and v.strip() == "")) for k, v in row_obj.items() if k != "row_index_excel"):
            continue
        roster.append(row_obj)

    missing = [k for k in expected if k not in column_map]

    # ID de template déterministe (nom fichier + feuille + header_row)
    template_id = f"tpl_{hex(abs(hash((name, ws.title, header_row_index))))[2:12]}"

    return {
        "template_id": template_id,
        "sheet_name": ws.title,
        "header_row_index": header_row_index,
        "column_map": column_map,
        "roster": roster,
        "missing_columns": missing,
    }
# ─────────────────────────── NEW: Timesheet Intake (pointage client)
@app.post("/timesheet-intake")
async def timesheet_intake(
    file_timesheet: UploadFile = File(...),
    holiday_dates: Optional[str] = Form(default=None),
    rules: Optional[str] = Form(default=None),
):
    fname = (getattr(file_timesheet, "filename", "") or "").lower()
    if not fname or not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
        logger.info(
            "timesheet-intake | rejected filename=%s mime=%s",
            getattr(file_timesheet, "filename", None),
            getattr(file_timesheet, "content_type", None),
        )
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm are supported")
    logger.info(
        "timesheet-intake | filename=%s mime=%s",
        file_timesheet.filename, getattr(file_timesheet, "content_type", None)
    )

    # Lecture + garde-fous
    try:
        content = await file_timesheet.read()
        if not content:
            raise HTTPException(status_code=400, detail="Empty file")
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail="File too large")
        wb = load_workbook(filename=BytesIO(content), data_only=True)

        # feuille + ligne d'entêtes robustes
        ws, header_row_index = _pick_best_sheet(wb)
        headers_dict = _extract_headers(ws, header_row_index)
        if not headers_dict:
            raise HTTPException(status_code=400, detail="No headers detected on the selected sheet.")
        logger.info(
            "timesheet-intake | sheet=%s header_row=%s headers_sample=%s",
            ws.title, header_row_index, list(headers_dict.values())[:10]
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel: {e}")

    rules_dict = _parse_rules(rules)
    holidays = set(_parse_holidays(holiday_dates))

    # Détection auto des colonnes (fuzzy)
    base_detected = _detect_columns(headers_dict)

    # Ponts d’alias
    def _resolve_detected(d: Dict[str, str]) -> Dict[str, str]:
        out = dict(d)
        alias_bridge = {
            "heures_norm": ["hrs_norm_010", "heures_norm_dec"],
            "hs_25": ["hs_25_020", "hs_25_dec"],
            "hs_50": ["hs_50_030", "hs_50_dec"],
            "hs_100": ["hs_100_050", "hs_100_dec"],
            "hs_feries": ["hrs_feries_091", "hs_feries_dec"],
            "date": ["date_debut"],
        }
        for target, aliases in alias_bridge.items():
            if target not in out:
                for a in aliases:
                    if a in out:
                        out[target] = out[a]
                        break
        return out

    detected = _resolve_detected(base_detected)

    # Si la date reste introuvable → heuristique
    if "date" not in detected and headers_dict:
        start_check = header_row_index + 1
        end_check = min(getattr(ws, "max_row", start_check - 1), start_check + 10)
        best_col, best_hits = None, 0
        for col_letter in headers_dict.keys():
            hits = 0
            for r in range(start_check, end_check + 1):
                try:
                    if _coerce_date(ws[f"{col_letter}{r}"].value):
                        hits += 1
                except Exception:
                    continue
            if hits > best_hits:
                best_hits, best_col = hits, col_letter
        sample_len = max(0, end_check - start_check + 1)
        if best_col and best_hits >= max(2, sample_len // 3):
            detected["date"] = best_col

    # Helpers locaux
    def _val_at(r: int, col_key: str):
        col_letter = detected.get(col_key)
        if not col_letter:
            return None
        try:
            return ws[f"{col_letter}{r}"].value
        except Exception:
            return None

    def _split_full_name(fn: Any) -> Tuple[Optional[str], Optional[str]]:
        """Heuristique : fichiers 'NOM PRENOM' ou 'Prenom Nom' → on sépare."""
        if not fn:
            return None, None
        s = str(fn).strip()
        if not s:
            return None, None
        parts = [p for p in re.split(r"\s+", s) if p]
        if len(parts) == 1:
            return parts[0], None
        prenom = parts[-1]
        nom = " ".join(parts[:-1])
        return nom, prenom

    # Preview (5 lignes), bornes sûres
    preview_rows: List[Dict[str, Any]] = []
    start = header_row_index + 1
    end = min(getattr(ws, "max_row", start - 1), start + 4)
    if end < start:
        return {
            "sheet_name": ws.title,
            "header_row_index": header_row_index,
            "detected_columns": detected,
            "preview_rows": [],
            "warnings": ["Aucune ligne de données trouvée sous la ligne d'entêtes."],
            "rules_used": rules_dict,
            "holiday_dates": list(holidays),
        }

    for r in range(start, end + 1):
        absence_raw = _val_at(r, "absence")

        demi_j = None
        if isinstance(absence_raw, str):
            ar = absence_raw.strip().lower()
            if (
                "demi" in ar
                or ar in {"am", "pm", "1/2", "0.5", "demi-j", "demi journee", "demi-journée"}
                or ar.replace(",", ".") == "0.5"
            ):
                demi_j = True

        # Date & fériés
        row_date = _coerce_date(_val_at(r, "date"))
        is_holiday = (row_date in holidays) if row_date else False

        # Identité avec fallback full_name → (nom, prenom)
        nom_v = _val_at(r, "nom")
        prenom_v = _val_at(r, "prenom")
        if (not nom_v or not prenom_v) and ("full_name" in detected):
            n, p = _split_full_name(_val_at(r, "full_name"))
            nom_v = nom_v or n
            prenom_v = prenom_v or p

        # Heures et jours (nb_jt)
        h_norm = _parse_hours_to_decimal(_val_at(r, "heures_norm"))
        nb_jt_val = _parse_days(_val_at(r, "nb_jt"))
        # Nouveaux compteurs de jours ↓↓↓
        nb_sans_solde_val = _parse_days(_val_at(r, "nb_sans_solde"))
        nb_jf_val = _parse_days(_val_at(r, "nb_jf"))
        # Conversions heures ↔ jours
        jours_calc: Optional[float] = None
        heures_from_days: Optional[float] = None

        if (h_norm is not None) and (nb_jt_val is None or nb_jt_val == 0):
            jours_calc = _hours_to_days(h_norm, rules_dict)

        if (nb_jt_val is not None) and (h_norm is None or h_norm == 0):
            heures_from_days = _days_to_hours(nb_jt_val, rules_dict)

        # Fallback demi-journée si rien de fourni
        if (h_norm is None) and (nb_jt_val is None) and demi_j:
            nb_jt_val = 0.5
            jours_calc = 0.5
            heures_from_days = _days_to_hours(0.5, rules_dict)

        preview_rows.append({
            "row_index_excel": r,
            "matricule": _val_at(r, "matricule"),
            "cin": _val_at(r, "cin"),
            "nom": nom_v,
            "prenom": prenom_v,
            "service": _val_at(r, "service"),
            "date": row_date,

            # Heures
            "heures_norm_dec": h_norm,
            "hs_25_dec": _parse_hours_to_decimal(_val_at(r, "hs_25")),
            "hs_50_dec": _parse_hours_to_decimal(_val_at(r, "hs_50")),
            "hs_100_dec": _parse_hours_to_decimal(_val_at(r, "hs_100")),
            "hs_feries_dec": _parse_hours_to_decimal(_val_at(r, "hs_feries")),

            # Jours
            "nb_jt": nb_jt_val,
            "nb_sans_solde": nb_sans_solde_val,   # <-- AJOUTÉ
            "nb_jf": nb_jf_val,                   # <-- AJOUTÉ

            # Conversions
            "jours_calc": jours_calc,
            "heures_from_days": heures_from_days,

            "demi_journee": demi_j,
            "is_holiday": is_holiday,
            "observations": _val_at(r, "observations"),
        })


    # Warnings
    warnings: List[str] = []
    if not detected:
        warnings.append("Aucune colonne n’a été reconnue automatiquement (en-têtes très atypiques).")
    for k in ("date", "heures_norm"):
        if k not in detected:
            warnings.append(f"Colonne importante non détectée : {k}")
    if detected.get("nom") and detected.get("prenom") and detected["nom"] == detected["prenom"]:
        warnings.append("Une seule colonne semble contenir NOM & PRÉNOM : les deux clés pointent sur la même colonne.")
    if "full_name" in detected and (("nom" not in detected) or ("prenom" not in detected)):
        warnings.append("Colonne 'full_name' détectée sans 'nom'/'prenom' dédiés : split heuristique appliqué dans l’aperçu.")

    def _column_has_hours(col_key: str) -> bool:
        col = detected.get(col_key)
        if not col:
            return False
        s0 = header_row_index + 1
        e0 = min(getattr(ws, "max_row", s0 - 1), s0 + 10)
        ok = 0
        for rr in range(s0, e0 + 1):
            try:
                if _parse_hours_to_decimal(ws[f"{col}{rr}"].value) is not None:
                    ok += 1
            except Exception:
                continue
        return ok >= 2

    for ck in ("heures_norm", "hs_25", "hs_50", "hs_100", "hs_feries"):
        if ck in detected and not _column_has_hours(ck):
            warnings.append(f"La colonne '{ck}' ne semble pas contenir des heures interprétables (échantillon).")

    return {
        "sheet_name": ws.title,
        "header_row_index": header_row_index,
        "detected_columns": detected,
        "preview_rows": preview_rows,
        "warnings": warnings,
        "rules_used": rules_dict,
        "holiday_dates": list(holidays),
    }


@app.post("/merge-intake")
async def merge_intake(request: Request):
    """
    Attend un JSON de la forme :
    {
      "template_roster": [ {...}, ... ],
      "timesheet_rows":  [ {...}, ... ],
      "timesheet_period": "YYYY-MM",
      "fuzzy_threshold_strict": 92,      # optionnel
      "fuzzy_threshold_loose": 85,       # optionnel
      "require_initial_match": true      # optionnel
    }
    """
    # 1) Lire le body brut et le parser nous-mêmes (pour éviter le 422 FastAPI)
    try:
        raw = await request.body()
        # tolérance encodage
        text = raw.decode("utf-8", errors="replace").strip()
        if not text:
            raise ValueError("empty body")
        payload = json.loads(text)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON body: {e}")

    # 2) Récupérer / normaliser les champs attendus
    def _coerce_list(v):
        if v is None:
            return []
        if isinstance(v, list):
            return v
        if isinstance(v, str):
            v = v.strip()
            if not v:
                return []
            try:
                parsed = json.loads(v)
                return parsed if isinstance(parsed, list) else []
            except Exception:
                return []
        return []

    template_roster = _coerce_list(payload.get("template_roster"))
    timesheet_rows  = _coerce_list(payload.get("timesheet_rows"))
    period         = payload.get("timesheet_period")

    # seuils / options avec défauts robustes
    def _to_int(v, d): 
        try: return int(v)
        except: return d
    def _to_bool(v, d):
        if isinstance(v, bool): return v
        if isinstance(v, str): return v.strip().lower() in {"1","true","yes","y"}
        return d

    strict  = _to_int(payload.get("fuzzy_threshold_strict"), 92)
    loose   = _to_int(payload.get("fuzzy_threshold_loose"), 85)
    require = _to_bool(payload.get("require_initial_match"), True)

    # garde-fous types
    if not isinstance(template_roster, list) or not isinstance(timesheet_rows, list):
        raise HTTPException(status_code=400, detail="template_roster and timesheet_rows must be lists")

    # 3) Appel du merge
    result = merge_rows(
        template_roster=template_roster,
        timesheet_rows=timesheet_rows,
        timesheet_period=period,
        strict=strict,
        loose=loose,
        require_initials=require,
    )
    return result

# ─────────────────────────── Export Excel (résultat merge) ───────────────────────────

@app.post("/merge-export")
async def merge_export(merge_result: Dict[str, Any] = Body(...)):
    """
    Deux cas possibles :

    1) Le front envoie directement le résultat de /merge-intake :
       {
         "matched_rows": [...],
         "missing_in_client": [...],
         "missing_in_roster": [...],
         "ambiguous": [...],
         "stats": {...},
         "duplicates_ref_keys": [...]
       }

    2) Le front envoie encore l'input brut :
       {
         "template_roster": [...],
         "timesheet_rows": [...],
         "timesheet_period": "YYYY-MM",
         "fuzzy_threshold_strict": 92,
         "fuzzy_threshold_loose": 85,
         "require_initial_match": true
       }

    Dans le cas 2), on refait le merge ici avant de générer l'Excel.
    """
    if merge_result is None:
        merge_result = {}

    try:
        logger.info("merge-export | raw_keys=%s", list(merge_result.keys()))
    except Exception:
        pass

    expected_keys = {
        "matched_rows",
        "missing_in_client",
        "missing_in_roster",
        "ambiguous",
        "stats",
    }

    # 1) Si on n'a PAS les clés de résultat mais qu'on a template_roster + timesheet_rows,
    #    on refait le merge nous-mêmes (cas où le front envoie encore l'input brut).
    if not any(k in merge_result for k in expected_keys) and \
       "template_roster" in merge_result and "timesheet_rows" in merge_result:
        try:
            logger.info("merge-export | detected raw input → recompute merge_rows")

            def _coerce_list(v):
                if v is None:
                    return []
                if isinstance(v, list):
                    return v
                if isinstance(v, str):
                    v = v.strip()
                    if not v:
                        return []
                    try:
                        parsed = json.loads(v)
                        return parsed if isinstance(parsed, list) else []
                    except Exception:
                        return []
                return []

            template_roster = _coerce_list(merge_result.get("template_roster"))
            timesheet_rows  = _coerce_list(merge_result.get("timesheet_rows"))
            period          = merge_result.get("timesheet_period")

            def _to_int(v, d):
                try:
                    return int(v)
                except Exception:
                    return d

            def _to_bool(v, d):
                if isinstance(v, bool):
                    return v
                if isinstance(v, str):
                    return v.strip().lower() in {"1", "true", "yes", "y"}
                return d

            strict  = _to_int(merge_result.get("fuzzy_threshold_strict"), 92)
            loose   = _to_int(merge_result.get("fuzzy_threshold_loose"), 85)
            require = _to_bool(merge_result.get("require_initial_match"), True)

            logger.info(
                "merge-export | recompute with counts template=%s timesheet=%s strict=%s loose=%s require=%s",
                len(template_roster), len(timesheet_rows), strict, loose, require
            )

            merge_result = merge_rows(
                template_roster=template_roster,
                timesheet_rows=timesheet_rows,
                timesheet_period=period,
                strict=strict,
                loose=loose,
                require_initials=require,
            )
            if period:
                merge_result["timesheet_period"] = period

            logger.info(
                "merge-export | recompute done: matched=%s missing_client=%s missing_roster=%s ambiguous=%s",
                len(merge_result.get("matched_rows") or []),
                len(merge_result.get("missing_in_client") or []),
                len(merge_result.get("missing_in_roster") or []),
                len(merge_result.get("ambiguous") or []),
            )

        except Exception as e:
            logger.error("merge-export | error while recomputing merge_rows: %s", e)

    # 2) À ce stade on est censés avoir un vrai résultat de merge_rows
    matched     = merge_result.get("matched_rows") or []
    miss_client = merge_result.get("missing_in_client") or []
    miss_roster = merge_result.get("missing_in_roster") or []
    ambiguous   = merge_result.get("ambiguous") or []
    stats       = merge_result.get("stats") or {}
    timesheet_period = merge_result.get("timesheet_period")


    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # ───────────── FEUILLE "Matched" AU FORMAT ROSTER PAIE ─────────────

    ws_matched = wb.create_sheet(title="Matched")

    # ---- Entête fixe (RH / Mois / Période du / Coefficient) ----

    # Ligne 1 : RH + nom client (si dispo dans stats)
    client_name = ""
    if isinstance(stats, dict):
        client_name = stats.get("client_name") or ""
    ws_matched["A1"] = "RH"
    ws_matched["B1"] = client_name

    # Ligne 2 : Mois + Année à partir de timesheet_period = "YYYY-MM"
    ws_matched["A2"] = "Mois"
    month_str = ""
    year_str = ""
    if isinstance(timesheet_period, str) and len(timesheet_period) >= 7:
        year_str = timesheet_period[0:4]
        month_str = timesheet_period[5:7]
    ws_matched["B2"] = month_str
    ws_matched["C2"] = year_str

    # Ligne 3 : Période du 01/MM/YYYY au dernier jour du mois
    ws_matched["A3"] = "Période du"
    if month_str and year_str:
        try:
            first_day = f"01/{month_str}/{year_str}"
            last_day_num = calendar.monthrange(int(year_str), int(month_str))[1]
            last_day = f"{last_day_num:02d}/{month_str}/{year_str}"
            ws_matched["B3"] = first_day
            ws_matched["C3"] = last_day
        except Exception:
            pass

    # Ligne 4 : Coefficient (optionnel, pris depuis stats si présent)
    ws_matched["A4"] = "Coefficient"
    if isinstance(stats, dict) and stats.get("coefficient") is not None:
        ws_matched["B4"] = stats["coefficient"]

    # (clé_dans_JSON, en-tête_excel)
    template_columns = [

        ("matricule_salarie",       "Matricule"),
        ("matricule_client",        "Matricule Client"),
        ("nombre",                  "Nombre"),
        ("nom",                     "Nom"),
        ("prenom",                  "Prénom"),
        ("num_contrat",             "N° Contrat"),
        ("num_avenant",            "N° Avenant"),
        ("date_debut",              "Date debut"),
        ("date_fin",                "Date fin"),
        ("service",                 "Service"),
        ("nb_jt",                   "NB JT"),
        ("nb_ji",                   "NB JI"),
        ("nb_cp_280",               "280 - NB CP"),
        ("nb_sans_solde",           "NB Sans Solde"),
        ("nb_jf",                   "NB JF"),
        ("tx_sal",                  "Tx Sal"),
        ("hrs_norm_010",            "010 - HRS NORM"),
        ("rappel_hrs_norm_140",     "140 - Rappel HRS NORM"),
        ("hs_25_020",               "020 - HS 25%"),
        ("hs_50_030",               "030 - HS 50%"),
        ("hs_100_050",              "050 - HS 100%"),
        ("hrs_feries_091",          "091 - HRS FERIES"),
        ("observations",            "Observations"),
        ("fin_mission",             "Fin Mission (Oui/Non)"),
    ]

    # Ligne d'en-têtes EXACTEMENT dans l'ordre du roster
    ws_matched.append([label for _, label in template_columns])

    def _get_val(row: Dict[str, Any], key: str):
        # Petits ponts pour les heures si besoin
        if key == "hrs_norm_010":
            return (
                row.get("hrs_norm_010")
                or row.get("heures_norm_dec")
                or row.get("heures_travaillees_decimal")
            )
        if key == "hs_25_020":
            return row.get("hs_25_020") or row.get("hs_25_dec")
        if key == "hs_50_030":
            return row.get("hs_50_030") or row.get("hs_50_dec")
        if key == "hs_100_050":
            return row.get("hs_100_050") or row.get("hs_100_dec")
        if key == "hrs_feries_091":
            return row.get("hrs_feries_091") or row.get("hs_feries_dec")
        return row.get(key, "")

    for row in matched:
        ws_matched.append([
            _get_val(row, key)
            for key, _ in template_columns
        ])

    # ───────────── AUTRES FEUILLES (debug générique) ─────────────

    other_sections = [
        ("Missing in Client", miss_client),
        ("Missing in Roster", miss_roster),
        ("Ambiguous",        ambiguous),
    ]

    for sheet_name, data in other_sections:
        ws = wb.create_sheet(title=sheet_name)
        if not isinstance(data, list) or not data:
            continue

        headers = sorted({
            k
            for row in data
            if isinstance(row, dict)
            for k in row.keys()
        })
        if not headers:
            continue

        ws.append(headers)
        for row in data:
            if not isinstance(row, dict):
                continue
            ws.append([row.get(h, "") for h in headers])

    # ───────────── FEUILLE STATS ─────────────

    ws_stats = wb.create_sheet(title="Stats")
    ws_stats.append(["metric", "value"])
    ws_stats.append(["matched_rows_count", len(matched)])
    ws_stats.append(["missing_in_client_count", len(miss_client)])
    ws_stats.append(["missing_in_roster_count", len(miss_roster)])
    ws_stats.append(["ambiguous_count", len(ambiguous)])

    if isinstance(stats, dict) and stats:
        ws_stats.append([])
        ws_stats.append(["stats_key", "stats_value"])
        for k, v in stats.items():
            ws_stats.append([str(k), str(v)])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="merge_export.xlsx"'
        },
    )
